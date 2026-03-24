import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
import io, time, warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

st.set_page_config(page_title="SBR Tracker Pipeline", page_icon="⚡", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
@import url("https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=Syne:wght@400;600;700;800&family=DM+Sans:wght@300;400;500&display=swap");
html,body,[class*="css"]{font-family:"DM Sans",sans-serif;background-color:#0d0f14;color:#e8e6e0}
.stApp{background:#0d0f14}
[data-testid="stSidebar"]{background:#111318!important;border-right:1px solid #1e2028!important}
[data-testid="stSidebar"] *{color:#c8c6c0!important}
[data-testid="stSidebar"] .stMarkdown h3{font-family:"Syne",sans-serif;font-size:10px;font-weight:700;letter-spacing:0.18em;text-transform:uppercase;color:#5a5e6e!important;margin-bottom:4px}
h1,h2,h3,h4{font-family:"Syne",sans-serif!important}
h1{font-size:2.1rem!important;font-weight:800!important;letter-spacing:-0.02em}
h2{font-size:1.1rem!important;font-weight:600!important;color:#9898a8!important}
.stButton>button{font-family:"Syne",sans-serif;font-weight:700;font-size:13px;letter-spacing:0.08em;text-transform:uppercase;background:linear-gradient(135deg,#4f7cff 0%,#7c4fff 100%);color:#fff;border:none;border-radius:6px;padding:10px 28px;width:100%;box-shadow:0 0 24px rgba(79,124,255,0.3);transition:all 0.2s}
.stButton>button:hover{transform:translateY(-1px);box-shadow:0 0 32px rgba(79,124,255,0.5)}
.stButton>button:disabled{background:#1e2028!important;color:#3a3d4a!important;box-shadow:none;transform:none}
.stDownloadButton>button{font-family:"Syne",sans-serif;font-weight:700;font-size:13px;letter-spacing:0.08em;text-transform:uppercase;background:linear-gradient(135deg,#00c896 0%,#00a0d4 100%);color:#fff;border:none;border-radius:6px;padding:10px 28px;width:100%;box-shadow:0 0 24px rgba(0,200,150,0.25);transition:all 0.2s}
.stDownloadButton>button:hover{transform:translateY(-1px);box-shadow:0 0 32px rgba(0,200,150,0.4)}
[data-testid="stFileUploader"]{background:#111318;border:1px solid #1e2028;border-radius:8px;padding:4px 8px}
hr{border-color:#1e2028!important}
[data-testid="stMetric"]{background:#111318;border:1px solid #1e2028;border-radius:10px;padding:16px 20px!important}
[data-testid="stMetricLabel"]{font-family:"DM Mono",monospace!important;font-size:10px!important;letter-spacing:0.12em;text-transform:uppercase;color:#5a5e6e!important}
[data-testid="stMetricValue"]{font-family:"Syne",sans-serif!important;font-size:2rem!important;font-weight:800!important;color:#e8e6e0!important}
[data-testid="stProgress"]>div>div{background:linear-gradient(90deg,#4f7cff,#7c4fff);border-radius:4px}
[data-testid="stProgress"]>div{background:#1e2028;border-radius:4px}
[data-testid="stExpander"]{background:#111318;border:1px solid #1e2028!important;border-radius:8px!important}
[data-testid="stExpander"] summary{color:#9898a8!important}
.mono{font-family:"DM Mono",monospace;font-size:11px;color:#5a5e6e}
</style>
""", unsafe_allow_html=True)

# ── CONSTANTS ──────────────────────────────────────────────────────────────
FMT_CURRENCY="$#,##0.00"; FMT_DATE="MM/DD/YYYY"; FMT_TEXT="@"; FMT_GENERAL="General"
COLUMN_FORMATS={1:FMT_GENERAL,2:FMT_GENERAL,3:FMT_TEXT,4:FMT_GENERAL,5:FMT_GENERAL,6:FMT_TEXT,7:FMT_TEXT,8:FMT_DATE,9:FMT_DATE,10:FMT_TEXT,11:FMT_CURRENCY,12:FMT_CURRENCY,13:FMT_CURRENCY,14:FMT_CURRENCY,15:FMT_CURRENCY,16:FMT_TEXT,17:FMT_TEXT,18:FMT_TEXT,19:FMT_TEXT,20:FMT_TEXT,21:FMT_TEXT,22:FMT_DATE,23:FMT_GENERAL,24:FMT_TEXT,25:FMT_TEXT,26:FMT_TEXT,27:FMT_TEXT,28:FMT_TEXT,29:FMT_DATE,30:FMT_TEXT,31:FMT_DATE,32:FMT_TEXT,33:FMT_TEXT,34:FMT_DATE,35:"mm/dd/yyyy",36:FMT_TEXT,37:FMT_TEXT}
MIN_VALID_DATE=pd.Timestamp("2000-01-01")
DOCLOG_STUDY_ID_IDX=6; DOCLOG_SUBTYPE_IDX=8; DOCLOG_DATE_ISSUED_IDX=18
HDR_STUDY_ID="Study Id"; HDR_BILL_SUB_DATE="Bill Submission Date"; HDR_LAG_TIME="Lag Time From Submission Date"
HDR_RESPONSE_TYPE="Response Type"; HDR_AB="Payment Received?"; HDR_AC="Last Payment Date"
HDR_AD="EOR/Objection Received?"; HDR_AE="Last EOR/Objection Date"
HDR_TIMELY="Timely Response (Response in 60 Days)"; HDR_AM="SBR Sent Status"
PROCESS_STATUSES={"bill resubmitted","sbr time lapsed","payment eor cases","ppo reduction","sbr sent after time lapsed","billing submission timeline expired","settled with agreement"}
PROTECTED_STATUSES={"sbr sent","study closed","sbr in queue","pending sbr"}
EOB_KEYWORDS=["eob zero","eob payment"]
VALID_SOL_STATUSES={"pending sbr","sbr sent","study closed"}

# ── HELPERS ────────────────────────────────────────────────────────────────
def normalize(v): return "" if v is None else str(v).strip().lower()
def safe_date(v):
    if v is None: return None
    try:
        if pd.isna(v): return None
    except: pass
    try: return pd.to_datetime(v).date()
    except: return None
def to_comparable_date(v):
    if v is None: return None
    s=str(v).strip()
    if s in ("","nan","NaT","None","NA","N/A"): return None
    try: return pd.to_datetime(s).date()
    except: return None
def safe_to_datetime(v):
    if v is None: return None
    if isinstance(v,(int,float)) and v==0: return None
    s=str(v).strip()
    if s in ("","0"): return None
    try:
        ts=pd.to_datetime(v)
        return None if ts<MIN_VALID_DATE else ts
    except: return None
def to_python_datetime(ts):
    if isinstance(ts,datetime): return datetime(ts.year,ts.month,ts.day)
    if isinstance(ts,date): return datetime(ts.year,ts.month,ts.day)
    return datetime(ts.year,ts.month,ts.day)
def to_date(v):
    if v is None: return None
    if isinstance(v,datetime): return v.date()
    if isinstance(v,date): return v
    try: return pd.to_datetime(v,format="mixed",dayfirst=False).date()
    except: return None
def to_date_obj(v):
    if v is None: return None
    if isinstance(v,datetime): return v.replace(hour=0,minute=0,second=0,microsecond=0)
    if isinstance(v,date): return datetime(v.year,v.month,v.day)
    try: return pd.to_datetime(v).to_pydatetime().replace(hour=0,minute=0,second=0,microsecond=0)
    except: return None
def to_number(v):
    try: return float(v)
    except: return None
def write_date_cell(cell,d):
    cell.value=datetime(d.year,d.month,d.day); cell.number_format="M/D/YYYY"; cell.alignment=Alignment(horizontal="center")
def contains_eob(z): return any(k in normalize(z) for k in EOB_KEYWORDS)
def is_paper_type(v): return normalize(v) in ("paper","paper attorney billed","paper employer billed")
def is_electronic_or_blank(v): return normalize(v) in ("","electronic")
def get_col_idx(headers,name):
    try: return headers.index(name)+1
    except ValueError: raise ValueError(f"Column \"{name}\" not found in sheet headers.")
def first_of_current_month():
    t=date.today(); return date(t.year,t.month,1)
def copy_font(f): return Font(name=f.name or "Calibri",size=f.size or 11,bold=f.bold,italic=f.italic,color=f.color.rgb if f.color and f.color.type=="rgb" else "000000")
def copy_fill(f):
    if f and f.fill_type and f.fill_type!="none":
        try: return PatternFill(fill_type=f.fill_type,start_color=f.start_color.rgb if f.start_color else "FFFFFF",end_color=f.end_color.rgb if f.end_color else "FFFFFF")
        except: pass
    return PatternFill(fill_type=None)
def copy_border(b):
    def s(x): return Side(style=x.style,color=x.color.rgb if x.color and x.color.type=="rgb" else "000000") if x and x.style else Side()
    return Border(left=s(b.left),right=s(b.right),top=s(b.top),bottom=s(b.bottom))
def copy_alignment(a): return Alignment(horizontal=a.horizontal or "center",vertical=a.vertical or "center",wrap_text=a.wrap_text)
def get_row_fmt(ws,row=2):
    out={}
    for c in range(1,38):
        cell=ws.cell(row=row,column=c)
        out[c]={"font":copy_font(cell.font),"fill":copy_fill(cell.fill),"border":copy_border(cell.border),"alignment":copy_alignment(cell.alignment),"number_format":cell.number_format}
    return out
def apply_fmt(ws,row,fmts):
    for c in range(1,38):
        cell=ws.cell(row=row,column=c); f=fmts.get(c,{})
        if f.get("font"): cell.font=f["font"]
        if f.get("fill"): cell.fill=f["fill"]
        if f.get("border"): cell.border=f["border"]
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=f.get("alignment",Alignment()).wrap_text)
        cell.number_format=COLUMN_FORMATS.get(c) or f.get("number_format",FMT_GENERAL)
def wb_to_bytes(wb):
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

# ── PHASE 1 ────────────────────────────────────────────────────────────────
def run_phase1(tracker_bytes,bdr_bytes,log):
    log("Loading Master Tracker and BDR...")
    tracker=pd.read_excel(io.BytesIO(tracker_bytes),sheet_name="Timeline")
    bdr=pd.read_csv(io.BytesIO(bdr_bytes),low_memory=False,dtype={"STUDY_ID":str,"CASEID":str})
    log(f"Tracker: {len(tracker):,} rows | BDR: {len(bdr):,} rows")
    existing=set(tracker["Study Id"].dropna().astype(str))
    bdr["_sid"]=bdr["STUDY_ID"].astype(str)
    new_cases=bdr[~bdr["_sid"].isin(existing)].copy(); new_cases.drop("_sid",axis=1,inplace=True)
    log(f"New cases found: {len(new_cases):,}")
    wb=load_workbook(io.BytesIO(tracker_bytes)); ws=wb["Timeline"]
    next_row=len(tracker)+2; fmts=get_row_fmt(ws,2)
    if len(new_cases)>0:
        for _,c in new_cases.iterrows():
            ws.cell(row=next_row,column=1).value=c["STUDY_ID"]; ws.cell(row=next_row,column=2).value=c["CASEID"]
            ws.cell(row=next_row,column=3).value=c["PATIENT_NAME"]; ws.cell(row=next_row,column=4).value=c["PATIENTREGID"]
            # EXTERNAL MRN — try exact name first, then case-insensitive fallback
            ext_mrn_val = ""
            if "EXTERNAL MRN" in c.index:
                ext_mrn_val = c["EXTERNAL MRN"] if pd.notna(c.get("EXTERNAL MRN")) else ""
            else:
                # fallback: search case-insensitively in case column name differs slightly
                for col_name in c.index:
                    if col_name.strip().upper() == "EXTERNAL MRN":
                        ext_mrn_val = c[col_name] if pd.notna(c[col_name]) else ""
                        break
            ws.cell(row=next_row,column=5).value=ext_mrn_val; ws.cell(row=next_row,column=6).value=c["BUSINESSNAME"]
            ws.cell(row=next_row,column=7).value=c["INSURANCE NAME"]
            ws.cell(row=next_row,column=8).value=pd.to_datetime(c["DATEOFSERVICE"]).date() if pd.notna(c["DATEOFSERVICE"]) else None
            ws.cell(row=next_row,column=9).value=pd.to_datetime(c["BILLDATE"]).date() if pd.notna(c["BILLDATE"]) else None
            next_row+=1
    bdr_lk=bdr.set_index("STUDY_ID").to_dict("index"); upd=0
    for r in range(2,next_row):
        sid=ws.cell(row=r,column=1).value
        if sid in bdr_lk:
            d=bdr_lk[sid]
            for col,key in [(10,"APPOINTMENTTYPE"),(11,"AMOUNTBILLED"),(12,"PAID"),(13,"WRITEOFF"),(14,"OUTSTANDING"),(15,"OMFS"),(16,"OWNER_NAME"),(17,"PORTFOLIO_NAME"),(18,"SUB OWNER_NAME"),(19,"STATUS")]:
                ws.cell(row=r,column=col).value=d.get(key,"")
            upd+=1
    if len(new_cases)>0:
        rs=len(tracker)+2; re=rs+len(new_cases)-1
        for r in range(rs,re+1):
            ws.cell(row=r,column=20).value="Yes"; ws.cell(row=r,column=21).value=None; ws.cell(row=r,column=22).value=None; ws.cell(row=r,column=23).value=None
            ws.cell(row=r,column=24).value=f'=IF(W{r}>=30,"30 days passed","Under 30 Days")'
            ws.cell(row=r,column=25).value=f'=IF(OR(AB{r}="Yes",AD{r}="Yes"),"Response Received","No Response")'
            ws.cell(row=r,column=26).value=None
            ws.cell(row=r,column=27).value=f'=IF(X{r}="Under 30 Days","No Action Required",IF(T{r}="No","No Action Required",IF(Y{r}="No Response","Send No Response Letter","Response Received")))'
            ws.cell(row=r,column=28).value="No"; ws.cell(row=r,column=29).value=None; ws.cell(row=r,column=30).value="No"
            ws.cell(row=r,column=31).value=None; ws.cell(row=r,column=32).value="No"; ws.cell(row=r,column=33).value="Letter not Sent"
            ws.cell(row=r,column=34).value=None
            ws.cell(row=r,column=35).value=f'=IF(Y{r}="Response Received",MAX(AE{r},AC{r})+60,DATE(1900,1,1))'
            ws.cell(row=r,column=36).value=f'=IF(Y{r}="Response Received","Yes","No")'
            ws.cell(row=r,column=37).value="Under Billing Cycle"
            apply_fmt(ws,r,fmts)
    closed=0
    for r in range(2,next_row):
        if ws.cell(row=r,column=19).value=="CLOSE": ws.cell(row=r,column=37).value="Study Closed"; closed+=1
    outstanding_issues=0
    for r in range(2,next_row):
        ov=ws.cell(row=r,column=14).value
        if ov and str(ov).strip() not in ("","0","0.0"):
            try:
                if float(ov)!=0: outstanding_issues+=1
            except: pass
    total=next_row-2; bdr_count=len(bdr)
    log(f"Row count {'MATCH' if bdr_count==total else f'MISMATCH (diff:{abs(bdr_count-total):,})'}: BDR={bdr_count:,} Tracker={total:,}")
    log(f"Updated J-S: {upd:,} | New: {len(new_cases):,} | Closed: {closed:,} | Outstanding issues: {outstanding_issues:,}")
    return wb_to_bytes(wb),{"new_cases":len(new_cases),"updated":upd,"closed":closed,"total_rows":total,"outstanding_issues":outstanding_issues}

# ── PHASE 2 ────────────────────────────────────────────────────────────────
def run_phase2(tracker_bytes,sub_bytes,sub_filename,log):
    log(f"Loading tracker and Submission Report ({sub_filename})...")
    tracker_df=pd.read_excel(io.BytesIO(tracker_bytes),sheet_name="Timeline")
    total=tracker_df.shape[0]; last=total+1
    if sub_filename.lower().endswith(".csv"):
        submission=pd.read_csv(io.BytesIO(sub_bytes),low_memory=False)
        log(f"Submission (CSV): {len(submission):,} rows")
    else:
        submission=pd.read_excel(io.BytesIO(sub_bytes))
        log(f"Submission (XLSX): {len(submission):,} rows")
    required=["STUDY_ID","EDI Service Type","Submission Date"]
    missing=[c for c in required if c not in submission.columns]
    if missing: raise ValueError(f"Submission Report missing columns: {missing}")
    lk=(submission.drop_duplicates(subset="STUDY_ID",keep="last").assign(STUDY_ID=lambda df:df["STUDY_ID"].astype(str).str.strip()).set_index("STUDY_ID")[["EDI Service Type","Submission Date"]].to_dict("index"))
    log(f"Lookup built: {len(lk):,} unique Study IDs")
    wb=load_workbook(io.BytesIO(tracker_bytes)); ws=wb["Timeline"]
    ws.insert_cols(23,3)
    pulled=not_found=0
    for r in range(2,last+1):
        sid=ws.cell(row=r,column=1).value
        if sid is None: continue
        sid=str(sid).strip()
        if sid in lk:
            ws.cell(row=r,column=23).value=lk[sid].get("EDI Service Type","")
            ws.cell(row=r,column=24).value=safe_date(lk[sid].get("Submission Date"))
            pulled+=1
        else: not_found+=1
    for r in range(2,last+1):
        tv=ws.cell(row=r,column=24).value
        if tv is not None and str(tv).strip()!="": ws.cell(row=r,column=25).value=f"=X{r}=V{r}"
    upd_uv=match_=na_=0
    for r in range(2,last+1):
        tw=ws.cell(row=r,column=23).value; tx=ws.cell(row=r,column=24).value
        if tw is None and tx is None: continue
        ev=ws.cell(row=r,column=22).value; es=str(ev).strip() if ev is not None else ""
        if es in ("nan","NaT","None","NA","N/A"): na_+=1
        elif ev is None or es=="" or to_comparable_date(tx)!=to_comparable_date(ev):
            ws.cell(row=r,column=21).value=tw; ws.cell(row=r,column=22).value=tx; upd_uv+=1
        else: match_+=1
    ws.delete_cols(23,3)
    lag=e_cnt=p_cnt=0
    for r in range(2,last+1):
        bst=ws.cell(row=r,column=21).value
        if normalize(bst)!="": ws.cell(row=r,column=23).value=f"=TODAY()-V{r}"; lag+=1
        if is_electronic_or_blank(bst): ws.cell(row=r,column=24).value=f'=IF(W{r}>=30,"30 days passed","Under 30 Days")'; e_cnt+=1
        elif is_paper_type(bst): ws.cell(row=r,column=24).value=f'=IF(W{r}>=45,"45 days passed","Under 45 Days")'; p_cnt+=1
    log(f"Pulled: {pulled:,} | Not found: {not_found:,} | U&V updated: {upd_uv:,} | NA skipped: {na_:,}")
    log(f"Lag rows: {lag:,} | 30-day: {e_cnt:,} | 45-day: {p_cnt:,}")
    return wb_to_bytes(wb),{"pulled":pulled,"not_found":not_found,"uv_updated":upd_uv,"lag_rows":lag,"e30":e_cnt,"p45":p_cnt}

# ── PHASE 3 ────────────────────────────────────────────────────────────────
def run_phase3(tracker_bytes,pay_bytes,pay_filename,log):
    log(f"Loading tracker and Payment Report ({pay_filename})...")
    tracker_df=pd.read_excel(io.BytesIO(tracker_bytes),sheet_name="Timeline")
    CHUNK=50_000; pay_lk={}
    if pay_filename.lower().endswith(".csv"):
        total_rows=0
        reader=pd.read_csv(io.BytesIO(pay_bytes),chunksize=CHUNK,low_memory=False,usecols=lambda c:c in ("STUDY_ID","PAYMENTDATE"))
        for chunk in reader:
            total_rows+=len(chunk); raw=chunk["PAYMENTDATE"]
            if pd.api.types.is_datetime64_any_dtype(raw): chunk["PAYMENTDATE"]=pd.to_datetime(raw,errors="coerce")
            elif pd.api.types.is_numeric_dtype(raw): chunk["PAYMENTDATE"]=pd.to_datetime(raw,unit="D",origin="1899-12-30",errors="coerce")
            else: chunk["PAYMENTDATE"]=pd.to_datetime(raw,errors="coerce")
            chunk.loc[chunk["PAYMENTDATE"].notna()&(chunk["PAYMENTDATE"]<MIN_VALID_DATE),"PAYMENTDATE"]=pd.NaT
            chunk=chunk.dropna(subset=["STUDY_ID","PAYMENTDATE"]); chunk["STUDY_ID"]=chunk["STUDY_ID"].astype(str).str.strip()
            for sid,dt in chunk.groupby("STUDY_ID")["PAYMENTDATE"].max().items():
                if sid not in pay_lk or dt>pay_lk[sid]: pay_lk[sid]=dt
        log(f"Payment (CSV): {total_rows:,} rows in {CHUNK:,}-row chunks")
    else:
        payment_df=pd.read_excel(io.BytesIO(pay_bytes),usecols=["STUDY_ID","PAYMENTDATE"]); raw=payment_df["PAYMENTDATE"]
        if pd.api.types.is_datetime64_any_dtype(raw): payment_df["PAYMENTDATE"]=pd.to_datetime(raw,errors="coerce")
        elif pd.api.types.is_numeric_dtype(raw): payment_df["PAYMENTDATE"]=pd.to_datetime(raw,unit="D",origin="1899-12-30",errors="coerce")
        else: payment_df["PAYMENTDATE"]=pd.to_datetime(raw,errors="coerce")
        payment_df.loc[payment_df["PAYMENTDATE"].notna()&(payment_df["PAYMENTDATE"]<MIN_VALID_DATE),"PAYMENTDATE"]=pd.NaT
        clean=payment_df.dropna(subset=["STUDY_ID","PAYMENTDATE"]).copy(); clean["STUDY_ID"]=clean["STUDY_ID"].astype(str).str.strip()
        pay_lk=clean.groupby("STUDY_ID")["PAYMENTDATE"].max().to_dict()
        log(f"Payment (XLSX): {len(payment_df):,} rows")
    if not pay_lk: raise ValueError("Payment Report produced empty pivot — check STUDY_ID and PAYMENTDATE columns.")
    log(f"Payment pivot: {len(pay_lk):,} unique Study IDs")
    wb=load_workbook(io.BytesIO(tracker_bytes)); ws=wb["Timeline"]; last=tracker_df.shape[0]+1
    CENTER=Alignment(horizontal="center",vertical="center")
    COL_STUDY_ID=1; COL_RESPONSE_TYPE=26; COL_PAYMENT_RECV=28; COL_PAYMENT_DATE=29
    matched=updated=already_yes=already_date_upd=already_date_kept=already_no_date=0
    blank_to_pay=combined=date_new=date_kept=not_found=0
    for r in range(2,last+1):
        sid_raw=ws.cell(row=r,column=COL_STUDY_ID).value
        if sid_raw is None: continue
        sid=str(sid_raw).strip()
        if sid not in pay_lk: not_found+=1; continue
        matched+=1; new_date=pay_lk[sid]
        pay_recv=ws.cell(row=r,column=COL_PAYMENT_RECV).value
        if str(pay_recv).strip().upper()!="NO":
            already_yes+=1
            existing_ts=safe_to_datetime(ws.cell(row=r,column=COL_PAYMENT_DATE).value)
            if existing_ts is not None:
                if new_date.date()>existing_ts.date():
                    ws.cell(row=r,column=COL_PAYMENT_DATE).value=to_python_datetime(new_date)
                    ws.cell(row=r,column=COL_PAYMENT_DATE).number_format="mm-dd-yy"
                    ws.cell(row=r,column=COL_PAYMENT_DATE).alignment=CENTER; already_date_upd+=1
                else:
                    ws.cell(row=r,column=COL_PAYMENT_DATE).value=to_python_datetime(existing_ts)
                    ws.cell(row=r,column=COL_PAYMENT_DATE).number_format="mm-dd-yy"
                    ws.cell(row=r,column=COL_PAYMENT_DATE).alignment=CENTER; already_date_kept+=1
            else: already_no_date+=1
            continue
        existing_resp=ws.cell(row=r,column=COL_RESPONSE_TYPE).value
        if not existing_resp or str(existing_resp).strip()=="": new_resp="Payment"; blank_to_pay+=1
        else: new_resp=f"Payment & {str(existing_resp).strip()}"; combined+=1
        ws.cell(row=r,column=COL_RESPONSE_TYPE).value=new_resp; ws.cell(row=r,column=COL_RESPONSE_TYPE).alignment=CENTER
        ws.cell(row=r,column=COL_PAYMENT_RECV).value="Yes"
        existing_ts=safe_to_datetime(ws.cell(row=r,column=COL_PAYMENT_DATE).value)
        if existing_ts is not None:
            if new_date.date()>existing_ts.date(): final_date=new_date; date_new+=1
            elif existing_ts.date()>new_date.date(): final_date=existing_ts; date_kept+=1
            else: final_date=existing_ts; date_kept+=1
        else: final_date=new_date; date_new+=1
        ws.cell(row=r,column=COL_PAYMENT_DATE).value=to_python_datetime(final_date)
        ws.cell(row=r,column=COL_PAYMENT_DATE).number_format="mm-dd-yy"
        ws.cell(row=r,column=COL_PAYMENT_DATE).alignment=CENTER; updated+=1
    log(f"Matched: {matched:,} | NO->YES: {updated:,} | Already YES: {already_yes:,} | Not found: {not_found:,}")
    log(f"Date refreshed: {already_date_upd:,} | Date cleaned: {already_date_kept:,} | Blank->Payment: {blank_to_pay:,} | Combined: {combined:,}")
    return wb_to_bytes(wb),{"matched":matched,"updated":updated,"already_yes":already_yes,"not_found":not_found,"date_new":date_new,"date_kept":date_kept}

# ── PHASE 4 ────────────────────────────────────────────────────────────────
def run_phase4(tracker_bytes,doclog_bytes,doclog_filename,log):
    log(f"Loading tracker and Doc-Log ({doclog_filename})...")
    tracker_df=pd.read_excel(io.BytesIO(tracker_bytes),sheet_name="Timeline")
    log(f"Tracker: {len(tracker_df):,} rows")
    if doclog_filename.lower().endswith(".csv"):
        doclog_df=pd.read_csv(io.BytesIO(doclog_bytes),low_memory=False,header=0,dtype=str); log(f"Doc-Log (CSV): {len(doclog_df):,} rows")
    else:
        doclog_df=pd.read_excel(io.BytesIO(doclog_bytes),header=0); log(f"Doc-Log (XLSX): {len(doclog_df):,} rows")
    if len(doclog_df.columns)<=DOCLOG_DATE_ISSUED_IDX: raise ValueError(f"Doc-Log needs at least {DOCLOG_DATE_ISSUED_IDX+1} columns, found {len(doclog_df.columns)}")
    cols=list(doclog_df.columns); c_sid=cols[DOCLOG_STUDY_ID_IDX]; c_sub=cols[DOCLOG_SUBTYPE_IDX]; c_dt=cols[DOCLOG_DATE_ISSUED_IDX]
    log(f"Doc-Log cols: Study ID=\"{c_sid}\" | SubType=\"{c_sub}\" | Date Issued=\"{c_dt}\"")
    doclog_df[c_sid]=doclog_df[c_sid].astype(str).str.strip()
    doclog_df[c_dt]=pd.to_datetime(doclog_df[c_dt],format="mixed",dayfirst=False,errors="coerce")
    deduped=(doclog_df.sort_values(c_dt,ascending=False,na_position="last").drop_duplicates(subset=c_sid,keep="first").set_index(c_sid))
    af_lk=deduped[c_sub].to_dict(); ag_dt_lk=deduped[c_dt].to_dict()
    ag_date_lk={k:v.date() if pd.notna(v) else None for k,v in ag_dt_lk.items()}
    log(f"Doc-Log lookup: {len(af_lk):,} unique Study IDs")
    wb=load_workbook(io.BytesIO(tracker_bytes)); ws=wb["Timeline"]
    headers=[ws.cell(row=1,column=c).value for c in range(1,ws.max_column+1)]
    COL_A=get_col_idx(headers,HDR_STUDY_ID); COL_V=get_col_idx(headers,HDR_BILL_SUB_DATE)
    COL_W=get_col_idx(headers,HDR_LAG_TIME); COL_Z=get_col_idx(headers,HDR_RESPONSE_TYPE)
    COL_AB=get_col_idx(headers,HDR_AB); COL_AC=get_col_idx(headers,HDR_AC)
    COL_AD=get_col_idx(headers,HDR_AD); COL_AE=get_col_idx(headers,HDR_AE)
    COL_TIMELY=get_col_idx(headers,HDR_TIMELY); COL_AM=get_col_idx(headers,HDR_AM)
    log(f"Columns: Z={COL_Z} AB={COL_AB} AC={COL_AC} AD={COL_AD} AE={COL_AE} Timely={COL_TIMELY} AM={COL_AM}")
    last=tracker_df.shape[0]+1; sol_cutoff=first_of_current_month()
    step26=step27=step28=not_in_doclog=step29=s29_no_date=s29_not_ubc=0
    s30_updated=s30_date_ok=s30_no_af=s30_protected=s30_not_scope=0
    s31_valid=s31_flagged=s31_past=s31_no_sol=step32=s32_skip=step34=s34_skip=0
    flagged_rows=[]
    log("Pass 1: Steps 26-30...")
    for r in range(2,last+1):
        sid_raw=ws.cell(row=r,column=COL_A).value
        if sid_raw is None: continue
        sid=str(sid_raw).strip()
        if sid in af_lk:
            af_val=af_lk[sid]; ag_ts=ag_dt_lk[sid]; af_str=str(af_val).strip() if af_val is not None else ""
            ag_do=to_date_obj(ag_ts) if pd.notna(ag_ts) else None
            if af_str:
                z_val=ws.cell(row=r,column=COL_Z).value; ad_val=ws.cell(row=r,column=COL_AD).value
                z_str=str(z_val).strip() if z_val else ""
                if normalize(ad_val)=="no" and z_str:
                    c=ws.cell(row=r,column=COL_Z); c.value=f"{z_str} & {af_str}"; c.alignment=Alignment(horizontal="center"); step26+=1
                elif not z_str:
                    c=ws.cell(row=r,column=COL_Z); c.value=af_str; c.alignment=Alignment(horizontal="center"); step27+=1
                ad_val=ws.cell(row=r,column=COL_AD).value
                if normalize(ad_val)=="no":
                    ws.cell(row=r,column=COL_AD).value="Yes"; dc=ws.cell(row=r,column=COL_AE)
                    dc.value=ag_do; dc.number_format="M/D/YYYY"; dc.alignment=Alignment(horizontal="center"); step28+=1
            else: not_in_doclog+=1
        else: not_in_doclog+=1
        am_val=ws.cell(row=r,column=COL_AM).value; am_norm=normalize(am_val)
        if am_norm=="under billing cycle":
            ab_v=ws.cell(row=r,column=COL_AB).value; ad_v=ws.cell(row=r,column=COL_AD).value
            ac_d=to_date(ws.cell(row=r,column=COL_AC).value); ae_d=to_date(ws.cell(row=r,column=COL_AE).value)
            resp=normalize(ab_v)=="yes" or normalize(ad_v)=="yes"
            if resp:
                vd=[d for d in [ac_d,ae_d] if d is not None]; sol_date=max(vd)+timedelta(days=60) if vd else None
            else: sol_date=None
            if sol_date is not None: ws.cell(row=r,column=COL_AM).value="PENDING SBR"; step29+=1
            else: s29_no_date+=1
        else: s29_not_ubc+=1
        am_val=ws.cell(row=r,column=COL_AM).value; am_norm=normalize(am_val)
        if am_norm in PROTECTED_STATUSES: s30_protected+=1; continue
        if am_norm not in PROCESS_STATUSES: s30_not_scope+=1; continue
        if sid not in af_lk: s30_no_af+=1; continue
        af30=af_lk[sid]; ag30=ag_date_lk[sid]
        if ag30 is None: s30_no_af+=1; continue
        ae_d=to_date(ws.cell(row=r,column=COL_AE).value)
        if ae_d is not None and ag30<=ae_d: s30_date_ok+=1; continue
        write_date_cell(ws.cell(row=r,column=COL_AE),ag30); ws.cell(row=r,column=COL_AM).value="PENDING SBR"
        z_v=ws.cell(row=r,column=COL_Z).value
        new_z=str(z_v).strip()+" & "+str(af30).strip() if normalize(z_v) and contains_eob(z_v) else str(af30).strip()
        ws.cell(row=r,column=COL_Z).value=new_z; s30_updated+=1
    log("Pass 2: Steps 31, 32, 34...")
    for r in range(2,last+1):
        ab_v=ws.cell(row=r,column=COL_AB).value; ad_v=ws.cell(row=r,column=COL_AD).value
        ac_d=to_date(ws.cell(row=r,column=COL_AC).value); ae_d=to_date(ws.cell(row=r,column=COL_AE).value)
        v_d=to_date(ws.cell(row=r,column=COL_V).value)
        resp=normalize(ab_v)=="yes" or normalize(ad_v)=="yes"
        if resp:
            vd=[d for d in [ac_d,ae_d] if d is not None]; sol_date=max(vd)+timedelta(days=60) if vd else None
        else: sol_date=None
        if sol_date is None: s31_no_sol+=1
        elif sol_date<sol_cutoff: s31_past+=1
        else:
            am_norm=normalize(ws.cell(row=r,column=COL_AM).value)
            if am_norm in VALID_SOL_STATUSES: s31_valid+=1
            else:
                row_data=[ws.cell(row=r,column=c).value for c in range(1,ws.max_column+1)]
                flagged_rows.append((r,row_data,sol_date,ws.cell(row=r,column=COL_AM).value)); s31_flagged+=1
        valid_resp=[d for d in [ac_d,ae_d] if d is not None]
        if v_d is not None and valid_resp:
            diff=(max(valid_resp)-v_d).days
            if 1<=diff<=60:
                ws.cell(row=r,column=COL_TIMELY).value="Yes"; ws.cell(row=r,column=COL_TIMELY).alignment=Alignment(horizontal="center"); step32+=1
            else: s32_skip+=1
        else: s32_skip+=1
        timely_val=ws.cell(row=r,column=COL_TIMELY).value
        if normalize(timely_val)!="no": s34_skip+=1; continue
        if not resp: s34_skip+=1; continue
        w_val=ws.cell(row=r,column=COL_W).value; w_num=to_number(w_val)
        if w_num is None and v_d is not None: w_num=(date.today()-v_d).days
        w_blank=w_val is None or (isinstance(w_val,str) and w_val.strip()=="")
        if not (w_blank or (w_num is not None and w_num<60)): s34_skip+=1; continue
        ws.cell(row=r,column=COL_TIMELY).value="Yes"; ws.cell(row=r,column=COL_TIMELY).alignment=Alignment(horizontal="center"); step34+=1
    if "Anomalies" in wb.sheetnames: del wb["Anomalies"]
    ws_anom=wb.create_sheet("Anomalies")
    hdr_font=Font(bold=True,color="FFFFFF"); hdr_fill=PatternFill("solid",fgColor="C00000")
    c_align=Alignment(horizontal="center",vertical="center"); flag_fill=PatternFill("solid",fgColor="FFE0E0")
    for ci,h in enumerate(headers,start=1):
        cell=ws_anom.cell(row=1,column=ci,value=h); cell.font=hdr_font; cell.fill=hdr_fill; cell.alignment=c_align
    for out_row,(ri,row_data,sol_dt,am_v) in enumerate(flagged_rows,start=2):
        for ci,val in enumerate(row_data,start=1):
            cell=ws_anom.cell(row=out_row,column=ci,value=val); cell.fill=flag_fill
            if isinstance(val,datetime): cell.number_format="M/D/YYYY"; cell.alignment=Alignment(horizontal="center")
    for ci in range(1,len(headers)+1):
        cl=get_column_letter(ci)
        max_len=max((len(str(ws_anom.cell(row=rw,column=ci).value or "")) for rw in range(1,len(flagged_rows)+2)),default=8)
        ws_anom.column_dimensions[cl].width=min(max_len+2,30)
    log(f"Steps 26-28: Z-appended={step26:,} Z-filled={step27:,} AD+AE={step28:,} Not-in-DocLog={not_in_doclog:,}")
    log(f"Step 29: PENDING SBR={step29:,} | No date={s29_no_date:,}")
    log(f"Step 30: Updated={s30_updated:,} | Date OK={s30_date_ok:,} | Protected={s30_protected:,}")
    log(f"Step 31 (cutoff {sol_cutoff.strftime('%m/%Y')}): Valid={s31_valid:,} | FLAGGED={s31_flagged:,} | Past={s31_past:,} | No SOL={s31_no_sol:,}")
    log(f"Step 32 Timely Yes={step32:,} | Step 34 Fixed={step34:,}")
    return wb_to_bytes(wb),{"s26":step26,"s27":step27,"s28":step28,"s29":step29,"s30":s30_updated,"s31_flagged":s31_flagged,"s32":step32,"s34":step34}

# ── SIDEBAR ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div style="padding:8px 0 20px"><div style="font-family:Syne,sans-serif;font-size:22px;font-weight:800;letter-spacing:-0.02em;color:#e8e6e0">SBR Pipeline</div><div style="font-family:DM Mono,monospace;font-size:10px;letter-spacing:0.14em;color:#3a3d4a;margin-top:2px">AUTOMATION SUITE v3.0</div></div>',unsafe_allow_html=True)
    st.markdown("### Phases to run")
    run_p1=st.checkbox("Phase 1 · Steps 1-6",value=True,help="New cases + BDR update")
    run_p2=st.checkbox("Phase 2 · Steps 7-15",value=True,help="Submission report + lag time")
    run_p3=st.checkbox("Phase 3 · Steps 16-22",value=True,help="Payment report processing")
    run_p4=st.checkbox("Phase 4 · Steps 24-34",value=True,help="Doc-Log + EOR + SBR + Timely Response + Anomalies sheet")
    st.markdown("---")
    st.markdown("### Options")
    keep_intermediate=st.checkbox("Keep intermediate files",value=False,help="Download each phase output separately")
    st.markdown("---")
    st.markdown('<div style="font-family:DM Mono,monospace;font-size:10px;color:#3a3d4a;line-height:2">ACCEPTED FORMATS<br>-- Master-Tracker .xlsx<br>-- BDR .csv<br>-- Submission .xlsx / .csv<br>-- Payment-Report .csv / .xlsx<br>-- Doc-Log .csv / .xlsx</div>',unsafe_allow_html=True)

# ── HEADER ───────────────────────────────────────────────────────────────────
col_title,col_badge=st.columns([5,1])
with col_title:
    st.markdown('<h1 style="margin-bottom:2px">SBR Tracker<br><span style="color:#4f7cff">Pipeline</span></h1><p style="font-family:DM Mono,monospace;font-size:11px;color:#3a3d4a;letter-spacing:0.1em;margin-top:4px">STEPS 1-34 · FULLY AUTOMATED</p>',unsafe_allow_html=True)
with col_badge:
    st.markdown('<div style="text-align:right;padding-top:16px"><div style="display:inline-block;background:#111318;border:1px solid #1e2028;border-radius:6px;padding:6px 12px;font-family:DM Mono,monospace;font-size:10px;color:#4f7cff;letter-spacing:0.1em">v3.0 READY</div></div>',unsafe_allow_html=True)

st.markdown("---")

# ── FILE UPLOADS ─────────────────────────────────────────────────────────────
st.markdown('<h2 style="font-family:Syne,sans-serif;font-size:13px;font-weight:700;letter-spacing:0.12em;text-transform:uppercase;color:#5a5e6e;margin-bottom:12px">01 - Upload Input Files</h2>',unsafe_allow_html=True)
col1,col2,col3=st.columns(3)
with col1:
    st.markdown('<div class="mono">Master Tracker · always required</div>',unsafe_allow_html=True)
    f_tracker=st.file_uploader("Master Tracker",type=["xlsx"],label_visibility="collapsed",key="tracker")
    st.markdown('<div class="mono" style="margin-top:10px">BDR Report · Phase 1</div>',unsafe_allow_html=True)
    f_bdr=st.file_uploader("BDR",type=["csv"],label_visibility="collapsed",key="bdr")
with col2:
    st.markdown('<div class="mono">Submission Report · Phase 2 · xlsx or csv</div>',unsafe_allow_html=True)
    f_sub=st.file_uploader("Submission",type=["xlsx","csv"],label_visibility="collapsed",key="sub")
    st.markdown('<div class="mono" style="margin-top:10px">Payment Report · Phase 3 · csv or xlsx</div>',unsafe_allow_html=True)
    f_pay=st.file_uploader("Payment Report",type=["csv","xlsx"],label_visibility="collapsed",key="pay")
with col3:
    st.markdown('<div class="mono">Doc-Log · Phase 4 · csv or xlsx</div>',unsafe_allow_html=True)
    f_doclog=st.file_uploader("Doc-Log",type=["csv","xlsx"],label_visibility="collapsed",key="doclog")
    st.markdown("<br>",unsafe_allow_html=True)
    file_map={"Master-Tracker":f_tracker,"BDR":f_bdr if run_p1 else True,"Submission":f_sub if run_p2 else True,"Payment Report":f_pay if run_p3 else True,"Doc-Log":f_doclog if run_p4 else True}
    status_html=""
    for fname,fobj in file_map.items():
        if fobj is True: status_html+=f'<div style="color:#3a3d4a;font-family:DM Mono,monospace;font-size:11px">- {fname} (skipped)</div>'
        elif fobj: status_html+=f'<div style="color:#00c896;font-family:DM Mono,monospace;font-size:11px">check {fname}</div>'
        else: status_html+=f'<div style="color:#ff6b6b;font-family:DM Mono,monospace;font-size:11px">x {fname}</div>'
    st.markdown(f'<div style="background:#111318;border:1px solid #1e2028;border-radius:8px;padding:12px 14px">{status_html}</div>',unsafe_allow_html=True)

st.markdown("---")

# ── PHASE CARDS ──────────────────────────────────────────────────────────────
st.markdown('<h2 style="font-family:Syne,sans-serif;font-size:13px;font-weight:700;letter-spacing:0.12em;text-transform:uppercase;color:#5a5e6e;margin-bottom:12px">02 - Pipeline Phases</h2>',unsafe_allow_html=True)
phase_cols=st.columns(4)
phase_meta=[("Phase 1","Steps 1-6","New cases + BDR update (cols J-S) + closed marking",run_p1),("Phase 2","Steps 7-15","Submission lookup · lag time · 30/45-day billing rules",run_p2),("Phase 3","Steps 16-22","Payment pivot · flag paid cases · date-only 3-way comparison",run_p3),("Phase 4","Steps 24-34","Doc-Log · EOR · PENDING SBR · Timely Response · Anomalies sheet",run_p4)]
for i,(ph,steps,desc,enabled) in enumerate(phase_meta):
    with phase_cols[i]:
        border="#4f7cff44" if enabled else "#1e2028"; alpha="1" if enabled else "0.45"
        st.markdown(f'<div style="background:#111318;border:1px solid {border};border-radius:10px;padding:16px;opacity:{alpha};min-height:120px"><div style="font-family:Syne,sans-serif;font-size:14px;font-weight:700;color:#e8e6e0">{ph}</div><div style="font-family:DM Mono,monospace;font-size:10px;color:#4f7cff;margin:2px 0 6px;letter-spacing:0.08em">{steps}</div><div style="font-size:11px;color:#5a5e6e;line-height:1.5">{desc}</div><div style="margin-top:8px;font-size:11px;color:{"#00c896" if enabled else "#3a3d4a"}">{"Enabled" if enabled else "Skipped"}</div></div>',unsafe_allow_html=True)

st.markdown("---")

# ── RUN BUTTON ───────────────────────────────────────────────────────────────
st.markdown('<h2 style="font-family:Syne,sans-serif;font-size:13px;font-weight:700;letter-spacing:0.12em;text-transform:uppercase;color:#5a5e6e;margin-bottom:12px">03 - Execute</h2>',unsafe_allow_html=True)
missing_files=[]
if not f_tracker: missing_files.append("Master-Tracker.xlsx")
if run_p1 and not f_bdr: missing_files.append("BDR.csv")
if run_p2 and not f_sub: missing_files.append("Submission (.xlsx or .csv)")
if run_p3 and not f_pay: missing_files.append("Payment Report (.csv or .xlsx)")
if run_p4 and not f_doclog: missing_files.append("Doc-Log (.csv or .xlsx)")
can_run=len(missing_files)==0
run_col,hint_col=st.columns([2,4])
with run_col:
    run_clicked=st.button("Run Pipeline" if can_run else "Missing required files",disabled=not can_run)
with hint_col:
    if missing_files: st.markdown(f'<div style="color:#ff6b6b;font-size:12px;padding-top:12px">Missing: {", ".join(missing_files)}</div>',unsafe_allow_html=True)
    else:
        phases_on=sum([run_p1,run_p2,run_p3,run_p4])
        st.markdown(f'<div style="color:#5a5e6e;font-size:12px;padding-top:12px">{phases_on} phase(s) enabled · all files ready</div>',unsafe_allow_html=True)

# ── EXECUTION ─────────────────────────────────────────────────────────────────
if run_clicked and can_run:
    st.markdown("---")
    st.markdown('<h2 style="font-family:Syne,sans-serif;font-size:13px;font-weight:700;letter-spacing:0.12em;text-transform:uppercase;color:#5a5e6e;margin-bottom:12px">04 - Running</h2>',unsafe_allow_html=True)
    log_lines=[]; all_stats={}; intermediate={}
    current_bytes=f_tracker.read(); overall_start=time.time(); had_error=False
    def add_log(msg,kind="def"): ts=datetime.now().strftime("%H:%M:%S"); log_lines.append((ts,msg,kind))
    prog_bar=st.progress(0); status_text=st.empty(); log_box=st.empty()
    def render_log():
        lines_html=""
        for ts,msg,kind in log_lines[-25:]:
            color={"ok":"#00c896","err":"#ff5f5f","warn":"#ffaa44","info":"#7a8aee"}.get(kind,"#6b6e7e")
            lines_html+=f'<div style="color:{color};font-family:DM Mono,monospace;font-size:11px;line-height:1.9"><span style="color:#3a3d4a">{ts}</span>  {msg}</div>'
        log_box.markdown(f'<div style="background:#080a0e;border:1px solid #1e2028;border-radius:8px;padding:14px 16px;max-height:260px;overflow-y:auto">{lines_html}</div>',unsafe_allow_html=True)
    # Use a mutable state dict to share data between helper functions.
    # This avoids the SyntaxError: "no binding for nonlocal" that occurs
    # when nested functions try to use nonlocal inside an if-block scope.
    total_phases = sum([run_p1, run_p2, run_p3, run_p4])
    state = {"done": 0, "error": False, "bytes": current_bytes}

    def lg(m): add_log(m); render_log()

    def run_one(n, label, fn, *args):
        if state["error"]: return
        try:
            status_text.markdown(
                f'<div style="font-family:DM Mono,monospace;font-size:12px;color:#4f7cff">Phase {n} - {label} running...</div>',
                unsafe_allow_html=True)
            add_log(f"=== Phase {n}: {label} ===", "info")
            result_bytes, stats = fn(*args, lg)
            state["bytes"] = result_bytes
            all_stats[f"phase{n}"] = stats
            if keep_intermediate: intermediate[f"Phase{n}"] = result_bytes
            add_log(f"Phase {n} complete", "ok")
            state["done"] += 1
            prog_bar.progress(state["done"] / total_phases)
            render_log()
        except Exception as e:
            add_log(f"Phase {n} ERROR: {e}", "err"); render_log()
            state["error"] = True

    def skip_one(n, label):
        add_log(f"Phase {n} ({label}) skipped", "warn"); render_log()

    if run_p1: run_one(1, "Steps 1-6",   run_phase1, state["bytes"], f_bdr.read())
    else:      skip_one(1, "Steps 1-6")

    if run_p2: run_one(2, "Steps 7-15",  run_phase2, state["bytes"], f_sub.read(), f_sub.name)
    else:      skip_one(2, "Steps 7-15")

    if run_p3: run_one(3, "Steps 16-22", run_phase3, state["bytes"], f_pay.read(), f_pay.name)
    else:      skip_one(3, "Steps 16-22")

    if run_p4: run_one(4, "Steps 24-34", run_phase4, state["bytes"], f_doclog.read(), f_doclog.name)
    else:      skip_one(4, "Steps 24-34")

    current_bytes = state["bytes"]
    had_error     = state["error"]

    elapsed=time.time()-overall_start; prog_bar.progress(1.0)
    if not had_error:
        status_text.markdown(f'<div style="font-family:DM Mono,monospace;font-size:12px;color:#00c896">Pipeline complete - {elapsed:.1f}s</div>',unsafe_allow_html=True)
        add_log(f"All phases done in {elapsed:.1f}s","ok"); render_log()
        st.markdown("---")
        st.markdown('<h2 style="font-family:Syne,sans-serif;font-size:13px;font-weight:700;letter-spacing:0.12em;text-transform:uppercase;color:#5a5e6e;margin-bottom:12px">05 - Results</h2>',unsafe_allow_html=True)
        s1=all_stats.get("phase1",{}); s3=all_stats.get("phase3",{}); s4=all_stats.get("phase4",{})
        m1,m2,m3,m4=st.columns(4)
        with m1: st.metric("Total Rows",f"{s1.get('total_rows',0):,}",delta=f"+{s1.get('new_cases',0):,} new")
        with m2: st.metric("New Cases Added",f"{s1.get('new_cases',0):,}")
        with m3: st.metric("Payments Updated",f"{s3.get('updated',0):,}")
        with m4: st.metric("Pending SBR",f"{s4.get('s29',0):,}")
        m5,m6,m7,m8=st.columns(4)
        with m5: st.metric("Timely Response Set",f"{s4.get('s32',0)+s4.get('s34',0):,}")
        with m6: st.metric("Anomalies Flagged",f"{s4.get('s31_flagged',0):,}",delta="see Anomalies sheet" if s4.get("s31_flagged",0)>0 else None)
        with m7: st.metric("EOR Updated",f"{s4.get('s28',0):,}")
        with m8: st.metric("Closed Cases",f"{s1.get('closed',0):,}")
        with st.expander("Detailed phase breakdown",expanded=False):
            for ph,stats in all_stats.items():
                st.markdown(f"**{ph.upper()}**")
                for k,v in stats.items():
                    st.markdown(f'<span style="font-family:DM Mono,monospace;font-size:11px;color:#7a7d8e">{k}</span>&nbsp;&nbsp;<span style="font-family:DM Mono,monospace;font-size:12px;color:#e8e6e0">{v:,}</span>',unsafe_allow_html=True)
                st.markdown("---")
        st.markdown("---")
        st.markdown('<h2 style="font-family:Syne,sans-serif;font-size:13px;font-weight:700;letter-spacing:0.12em;text-transform:uppercase;color:#5a5e6e;margin-bottom:12px">06 - Download Output</h2>',unsafe_allow_html=True)
        ts_str=datetime.now().strftime("%Y%m%d_%H%M%S"); out_filename=f"Master-Tracker_Updated_{ts_str}.xlsx"
        dl_col,info_col=st.columns([2,4])
        with dl_col:
            st.download_button(label="Download Updated Tracker",data=current_bytes,file_name=out_filename,mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with info_col:
            anom_note=" · Anomalies sheet included" if s4.get("s31_flagged",0)>0 else ""
            st.markdown(f'<div style="background:#111318;border:1px solid #1e2028;border-radius:8px;padding:14px 18px;margin-top:4px"><div style="font-family:DM Mono,monospace;font-size:11px;color:#5a5e6e">OUTPUT FILE</div><div style="font-family:DM Mono,monospace;font-size:13px;color:#e8e6e0;margin-top:4px">{out_filename}</div><div style="font-family:DM Mono,monospace;font-size:10px;color:#3a3d4a;margin-top:6px">Phases: {", ".join(all_stats.keys())} · Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}{anom_note}</div></div>',unsafe_allow_html=True)
        if keep_intermediate and intermediate:
            st.markdown("**Intermediate files:**")
            int_cols=st.columns(len(intermediate))
            for i,(ph,data) in enumerate(intermediate.items()):
                with int_cols[i]:
                    st.download_button(label=f"Download {ph}",data=data,file_name=f"Master-Tracker_{ph}_{ts_str}.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",key=f"dl_{ph}")
    else:
        status_text.markdown('<div style="font-family:DM Mono,monospace;font-size:12px;color:#ff5f5f">Pipeline stopped - check log above</div>',unsafe_allow_html=True)

st.markdown("---")
st.markdown('<div style="font-family:DM Mono,monospace;font-size:10px;color:#2a2d36;text-align:center;padding:8px 0">SBR TRACKER PIPELINE · STEPS 1-34 · FOR INTERNAL USE</div>',unsafe_allow_html=True)
